"""
============================================================
File:
    vobes_workbook.py

Project:
    VOBES Migration Tool

Purpose:
    Excel workbook access layer for the VOBES migration tools.

Description:
    Provides a Python abstraction for loading the VOBES
    workbook, reading and updating worksheet content, and
    accessing named ranges used by the migration workflow.

Primary Class:
    VobesWorkbook

Dependencies:
    openpyxl
    pinchar_engine.py

Notes:
    This module isolates Excel-specific operations from the
    service and GUI layers.

Last Updated:
    2026-07-31

============================================================
"""
from openpyxl import load_workbook
from pinchar_engine import PincharEngine


class VobesWorkbook:

    def __init__(self, filename):

        self.filename = str(filename)

        self.wb = load_workbook(
            self.filename,
            keep_vba=True,
            data_only=False
        )

        self.ws = self.wb["Formblatt"]
        
        self.pinchar = PincharEngine(self)

    # ==================================================
    # Named Range Support
    # ==================================================

    def get_named_cell(self, name):

        dn = self.wb.defined_names[name]

        sheet_name, cell_addr = \
            list(dn.destinations)[0]

        ws = self.wb[sheet_name]

        return ws[cell_addr]

    def get_named_value(self, name):

        return self.get_named_cell(
            name
        ).value

    def set_named_value(
            self,
            name,
            value):

        self.get_named_cell(
            name
        ).value = value

    # ==================================================
    # General Workbook Info
    # ==================================================

    def get_part_number(self):

        return self.get_named_value(
            "Teilenummer"
        )

    def get_pin_count(self):

        return self.get_named_value(
            "SumPins"
        )

    def get_slot_count(self):

        return self.get_named_value(
            "SumSlots"
        )

    def get_status(self):

        return self.get_named_value(
            "State"
        )

    def set_status(self, text):

        self.set_named_value(
            "State",
            text
        )

    def get_data_checked(self):

        return self.get_named_value(
            "DataChecked"
        )

    def set_data_checked(self, text):

        self.set_named_value(
            "DataChecked",
            text
        )

    # ==================================================
    # Pin Area
    # ==================================================

    def get_pin_bounds(self):

        upper = self.get_named_cell(
            "PinsUpper"
        ).row

        lower = self.get_named_cell(
            "PinsLower"
        ).row

        return (
            upper + 1,
            lower - 1
        )

    def get_pins(self):

        start_row, end_row = \
            self.get_pin_bounds()

        pins = []

        for row in range(
                start_row,
                end_row + 1):

            pin_type = self.ws.cell(
                row=row,
                column=52
            ).value
            pinchar_definition = \
                self.pinchar.get_definition(
                    pin_type
                )
            
            #
            # Debug first pin only
            #

            pins.append({

                "row": row,

                "slot":
                    self.ws.cell(
                        row=row,
                        column=2
                    ).value,

                "pin":
                    self.ws.cell(
                        row=row,
                        column=7
                    ).value,

                "description":
                    self.ws.cell(
                        row=row,
                        column=15
                    ).value,

                "checksum":
                    self.ws.cell(
                        row=row,
                        column=33
                    ).value,

                "pin_type": 
                    self.ws.cell(
                        row=row,
                        column=52
                    ).value,
                
                "characteristic":
                    self.ws.cell(
                        row=row,
                        column=65
                    ).value,

                #
                # Current fields
                #

                "i1":
                    self.ws.cell(
                        row=row,
                        column=78
                    ).value,

                "i2":
                    self.ws.cell(
                        row=row,
                        column=85
                    ).value,

                "i3":
                    self.ws.cell(
                        row=row,
                        column=92
                    ).value,

                "i4":
                    self.ws.cell(
                        row=row,
                        column=99
                    ).value,

                "i5":
                    self.ws.cell(
                        row=row,
                        column=106
                    ).value,
                    
                "i6":
                    self.ws.cell(
                        row=row,
                        column=113
                    ).value,
                
                "t1":
                    self.ws.cell(
                        row=row,
                        column=120
                    ).value,
                
                "t2":
                    self.ws.cell(
                        row=row,
                        column=127
                    ).value,

                #
                # CL Info
                #

                "clamp":
                    self.ws.cell(
                        row=row,
                        column=136
                    ).value,

                "function":
                    self.ws.cell(
                        row=row,
                        column=143
                    ).value,

                "direction":
                    self.ws.cell(
                        row=row,
                        column=158
                    ).value,

                "voltage":
                    self.ws.cell(
                        row=row,
                        column=163
                    ).value,

                "utilization":
                    self.ws.cell(
                        row=row,
                        column=171
                    ).value,

                "signal_type":
                    self.ws.cell(
                        row=row,
                        column=176
                    ).value,
                                    
                "pinchar":
                    pinchar_definition,
            })

        return pins
    
    def debug_pin_row(self, row):

        print()
        print("=" * 80)
        print("ROW", row)
        print("=" * 80)

        for col in range(1, 240):
            value = self.ws.cell(
                row=row,
                column=col
            ).value
            if value not in (None, ""):
                print(
                    f"COL {col:3d} = {value}"
                )

    def get_pin_by_name(
            self,
            slot_name,
            pin_number):

        for pin in self.get_pins():

            if (
                str(pin["slot"]) ==
                str(slot_name)
                and
                str(pin["pin"]) ==
                str(pin_number)
            ):

                return pin

        return None

    # ==================================================
    # Slot Area
    # ==================================================

    def get_slot_bounds(self):

        upper = self.get_named_cell(
            "SlotsUpper"
        ).row

        lower = self.get_named_cell(
            "SlotsLower"
        ).row

        return (
            upper + 1,
            lower - 1
        )

    def get_slots(self):

        start_row, end_row = \
            self.get_slot_bounds()

        slots = []

        for row in range(
                start_row,
                end_row + 1):

            slots.append({

                "row":
                    row,

                "name":
                    self.ws.cell(
                        row=row,
                        column=2
                    ).value
            })

        return slots

    # ==================================================
    # Direct Cell Access
    # ==================================================

    def get_cell(
            self,
            row,
            column):

        return self.ws.cell(
            row=row,
            column=column
        ).value

    def set_cell(
            self,
            row,
            column,
            value):

        self.ws.cell(
            row=row,
            column=column
        ).value = value

    # ==================================================
    # Pin Editing
    # ==================================================

    def set_pin_description(
            self,
            row,
            text):

        self.ws.cell(
            row=row,
            column=15
        ).value = text

    def set_pin_type(
            self,
            row,
            text):

        self.ws.cell(
            row=row,
            column=52
        ).value = text
        
    def set_pin_characteristic(
            self,
            row,
            value):
        self.ws.cell(
            row=row,
            column=65
        ).value = value

    def set_pin_clamp(
            self,
            row,
            value):
        self.ws.cell(
            row=row,
            column=136
        ).value = value

    def set_pin_function(
            self,
            row,
            value):
        self.ws.cell(
            row=row,
            column=143
        ).value = value
    
    def set_pin_direction(
            self,
            row,
            value):
        self.ws.cell(
            row=row,
            column=158
        ).value = value
        
    def set_pin_voltage(
            self,
            row,
            value):
        self.ws.cell(
            row=row,
            column=163
        ).value = value
        
    def set_pin_utilization(
            self,
            row,
            value):
        self.ws.cell(
            row=row,
            column=171
        ).value = value
        
    def set_pin_i1(self, row, val):
        self.ws.cell(row=row, column=78).value = val
    
    def set_pin_i2(self, row, val):
        self.ws.cell(row=row, column=85).value = val
        
    def set_pin_i3(self, row, val):
        self.ws.cell(row=row, column=92).value = val
        
    def set_pin_i4(self, row, val):
        self.ws.cell(row=row, column=99).value = val
        
    def set_pin_i5(self, row, val):
        self.ws.cell(row=row, column=106).value = val
        
    def set_pin_i6(self, row, val):
        self.ws.cell(row=row, column=113).value = val
    
    def set_pin_t1(self, row, val):
        self.ws.cell(row=row, column=120).value = val
    
    def set_pin_t2(self, row, val):
        self.ws.cell(row=row, column=127).value = val
        

    # ==================================================
    # Utility
    # ==================================================

    def print_summary(self):

        print()

        print("Part Number :",
              self.get_part_number())

        print("Pins        :",
              self.get_pin_count())

        print("Slots       :",
              self.get_slot_count())

        print("State       :",
              self.get_status())

        print("Checked     :",
              self.get_data_checked())

        print()

    # ==================================================
    # Save
    # ==================================================

    def save(self, filename=None):

        if filename is None:

            filename = self.filename

        self.wb.save(filename)